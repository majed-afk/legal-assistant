#!/usr/bin/env python3
"""
تصحيح 208 إجابة قانونية تلقائياً باستخدام Claude + النصوص الرسمية
Corrects all 208 Q&A answers using Claude API and official law text.

Usage:
    python -m backend.tools.correct_qa
    python -m backend.tools.correct_qa --start 50  # Resume from Q&A #50
    python -m backend.tools.correct_qa --ids 30,46,52  # Correct specific IDs only
"""

import json
import os
import re
import sys
import time
import argparse
import openpyxl
import anthropic
from pathlib import Path
from dotenv import load_dotenv

# Load env
ROOT = Path(__file__).resolve().parent.parent.parent
load_dotenv(ROOT / ".env")

ARTICLES_PATH = ROOT / "backend" / "data" / "ahwal_clean_articles.json"
QA_EXCEL_PATH = ROOT.parent / "Downloads" / "legal_qa_208_complete.xlsx"
OUTPUT_PATH = ROOT / "backend" / "data" / "corrected_qa.json"
PROGRESS_PATH = ROOT / "backend" / "data" / "correction_progress.json"


def load_articles():
    """Load all clean articles indexed by article number."""
    with open(ARTICLES_PATH, "r", encoding="utf-8") as f:
        articles = json.load(f)
    index = {}
    for a in articles:
        index[a["article_number"]] = a
    return index


def load_qa_from_excel():
    """Load 208 Q&A entries from Excel file."""
    wb = openpyxl.load_workbook(QA_EXCEL_PATH, read_only=True)
    ws = wb.active
    entries = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        entry = {
            "id": int(row[0]),
            "category": row[1] or "",
            "chapter": row[2] or "",
            "section": row[3] or "",
            "question_formal": row[4] or "",
            "question_colloquial": row[5] or "",
            "original_answer": row[6] or "",
            "cited_articles_raw": row[7] or "",
        }
        entries.append(entry)
    wb.close()
    return entries


def parse_article_numbers(raw: str) -> list[int]:
    """Extract article numbers from المواد column."""
    numbers = re.findall(r'(\d+)', str(raw))
    return [int(n) for n in numbers]


def get_relevant_articles(article_nums: list[int], articles_index: dict, context_range: int = 3) -> str:
    """Get text of cited articles + nearby articles for context."""
    all_nums = set()
    for num in article_nums:
        # Add the cited article + nearby articles
        for i in range(max(1, num - context_range), num + context_range + 1):
            if i in articles_index:
                all_nums.add(i)

    result_parts = []
    for num in sorted(all_nums):
        a = articles_index[num]
        marker = " ⬅️ [مُستشهد بها]" if num in article_nums else ""
        result_parts.append(
            f"المادة ({num}){marker} — {a['topic']} — {a['chapter']} > {a['section']}:\n{a['text']}"
        )
    return "\n\n".join(result_parts)


CORRECTION_PROMPT = """أنت مدقق قانوني متخصص في نظام الأحوال الشخصية السعودي.

مهمتك: تصحيح الإجابة أدناه بناءً على النصوص الرسمية المرفقة.

## السؤال (فصحى):
{question_formal}

## السؤال (عامي):
{question_colloquial}

## الإجابة الأصلية (تحتاج تصحيح):
{original_answer}

## المواد المُستشهد بها أصلاً:
{cited_articles_raw}

## النصوص الرسمية للمواد (المصدر الوحيد للحقيقة):
{articles_text}

## تعليمات التصحيح:
1. **تحقق من أرقام المواد**: إذا كان رقم المادة المذكور في الإجابة خاطئاً، صححه للرقم الصحيح من النصوص أعلاه.
2. **تحقق من الاقتباسات**: أي نص بين علامتي تنصيص يجب أن يكون مطابقاً حرفياً للنص الرسمي. صحح أي اقتباس محرّف.
3. **تحقق من المعلومات القانونية**: المدد، الأعمار، الحقوق، الشروط — يجب أن تطابق النص الرسمي تماماً.
4. **حدد المواد الصحيحة**: إذا كانت المواد المُستشهد بها أصلاً خاطئة، حدد المواد الصحيحة من النصوص المرفقة.
5. **لا تخترع**: لا تضف معلومات غير موجودة في النصوص الرسمية.

## شكل الإجابة المطلوب (JSON):
أجب بـ JSON فقط بهذا الشكل:
{{
    "corrected_answer": "الإجابة المصححة كاملة بنفس الأسلوب (📖 السند النظامي + 📋 التفصيل + 📌 ملاحظات + ⚠️ تحذيرات)",
    "corrected_articles": ["المادة X", "المادة Y"],
    "changes_made": ["وصف مختصر لكل تغيير تم"],
    "severity": "none|minor|major|critical"
}}

أجب بـ JSON فقط بدون أي نص إضافي."""


def correct_single_qa(client, entry: dict, articles_index: dict) -> dict:
    """Correct a single Q&A entry using Claude API."""
    # Parse cited articles
    cited_nums = parse_article_numbers(entry["cited_articles_raw"])

    # Get relevant article texts (cited + nearby for context)
    articles_text = get_relevant_articles(cited_nums, articles_index, context_range=5)

    # Build prompt
    prompt = CORRECTION_PROMPT.format(
        question_formal=entry["question_formal"],
        question_colloquial=entry["question_colloquial"],
        original_answer=entry["original_answer"],
        cited_articles_raw=entry["cited_articles_raw"],
        articles_text=articles_text,
    )

    # Call Claude
    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=2000,
        messages=[{"role": "user", "content": prompt}],
    )

    raw_text = response.content[0].text.strip()

    # Parse JSON from response
    # Handle markdown code blocks
    if "```json" in raw_text:
        raw_text = raw_text.split("```json")[1].split("```")[0].strip()
    elif "```" in raw_text:
        raw_text = raw_text.split("```")[1].split("```")[0].strip()

    try:
        result = json.loads(raw_text)
    except json.JSONDecodeError:
        result = {
            "corrected_answer": raw_text,
            "corrected_articles": [f"المادة {n}" for n in cited_nums],
            "changes_made": ["فشل تحليل JSON — الإجابة الخام محفوظة"],
            "severity": "unknown",
        }

    return {
        "id": entry["id"],
        "category": entry["category"],
        "question_formal": entry["question_formal"],
        "question_colloquial": entry["question_colloquial"],
        "original_answer": entry["original_answer"],
        "original_articles": entry["cited_articles_raw"],
        "corrected_answer": result.get("corrected_answer", ""),
        "corrected_articles": result.get("corrected_articles", []),
        "changes_made": result.get("changes_made", []),
        "severity": result.get("severity", "unknown"),
    }


def load_progress() -> dict:
    """Load correction progress."""
    if PROGRESS_PATH.exists():
        with open(PROGRESS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"completed_ids": [], "results": []}


def save_progress(progress: dict):
    """Save correction progress."""
    with open(PROGRESS_PATH, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


def main():
    parser = argparse.ArgumentParser(description="Correct 208 Q&A entries")
    parser.add_argument("--start", type=int, default=1, help="Start from Q&A ID")
    parser.add_argument("--ids", type=str, default="", help="Comma-separated IDs to correct")
    parser.add_argument("--resume", action="store_true", help="Resume from last progress")
    args = parser.parse_args()

    # Load data
    print("📚 تحميل المواد القانونية...")
    articles_index = load_articles()
    print(f"   ✅ {len(articles_index)} مادة")

    print("📋 تحميل الأسئلة والأجوبة...")
    qa_entries = load_qa_from_excel()
    print(f"   ✅ {len(qa_entries)} سؤال")

    # Filter entries
    if args.ids:
        target_ids = [int(x.strip()) for x in args.ids.split(",")]
        qa_entries = [e for e in qa_entries if e["id"] in target_ids]
        print(f"   🎯 تصحيح {len(qa_entries)} سؤال محدد")
    elif args.start > 1:
        qa_entries = [e for e in qa_entries if e["id"] >= args.start]
        print(f"   🎯 بدء من السؤال #{args.start}")

    # Load progress
    progress = load_progress() if args.resume else {"completed_ids": [], "results": []}
    completed_ids = set(progress["completed_ids"])

    # Skip already completed
    if args.resume:
        qa_entries = [e for e in qa_entries if e["id"] not in completed_ids]
        print(f"   ⏭️ تخطي {len(completed_ids)} سؤال مكتمل، متبقي {len(qa_entries)}")

    # Initialize Claude client
    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key:
        print("❌ ANTHROPIC_API_KEY غير موجود في .env")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)

    # Process each entry
    stats = {"none": 0, "minor": 0, "major": 0, "critical": 0, "unknown": 0}
    total = len(qa_entries)

    print(f"\n🔧 بدء التصحيح ({total} سؤال)...\n")

    for i, entry in enumerate(qa_entries):
        qid = entry["id"]
        print(f"  [{i+1}/{total}] سؤال #{qid}: {entry['question_formal'][:60]}...", end=" ", flush=True)

        try:
            result = correct_single_qa(client, entry, articles_index)
            severity = result["severity"]
            stats[severity] = stats.get(severity, 0) + 1

            icon = {"none": "✅", "minor": "🔵", "major": "🟡", "critical": "🔴"}.get(severity, "⚪")
            changes = len(result["changes_made"])
            print(f"{icon} {severity} ({changes} تغيير)")

            # Save progress
            progress["results"].append(result)
            progress["completed_ids"].append(qid)
            save_progress(progress)

            # Rate limiting
            time.sleep(0.5)

        except anthropic.RateLimitError:
            print("⏳ rate limit — انتظار 30 ثانية...")
            time.sleep(30)
            # Retry
            try:
                result = correct_single_qa(client, entry, articles_index)
                progress["results"].append(result)
                progress["completed_ids"].append(qid)
                save_progress(progress)
                print(f"  ✅ نجح بعد الانتظار")
            except Exception as e2:
                print(f"  ❌ فشل: {e2}")

        except Exception as e:
            print(f"❌ خطأ: {e}")
            continue

    # Save final output
    # Merge with any previous results
    all_results = {r["id"]: r for r in progress["results"]}
    final_results = sorted(all_results.values(), key=lambda x: x["id"])

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(final_results, f, ensure_ascii=False, indent=2)

    # Print summary
    print(f"\n{'='*60}")
    print(f"📊 ملخص التصحيح:")
    print(f"   ✅ بدون تغيير: {stats['none']}")
    print(f"   🔵 تغييرات طفيفة: {stats['minor']}")
    print(f"   🟡 تغييرات جوهرية: {stats['major']}")
    print(f"   🔴 أخطاء حرجة: {stats['critical']}")
    print(f"   📁 محفوظ في: {OUTPUT_PATH}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
